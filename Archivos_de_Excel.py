from openpyxl import Workbook, load_workbook

# Crear un archivo de Excel y escribir datos
wb = Workbook()
hoja = wb.active
hoja.title = "Estudiantes"
hoja["A1"], hoja["B1"] = "ID", "Nombre"  # Encabezados de las columnas
hoja.append([1, "Juan"])
hoja.append([2, "Ana"])
wb.save("archivo_excel.xlsx")  # Guardar como .xlsx

print("Archivo de Excel creado exitosamente como 'archivo_excel.xlsx'.")

# Leer datos del archivo de Excel
wb = load_workbook("archivo_excel.xlsx")
hoja = wb["Estudiantes"]
for fila in hoja.iter_rows(values_only=True):  # Leer fila por fila
    print(fila)
